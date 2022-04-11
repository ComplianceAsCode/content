#!/usr/bin/env python3

import argparse
import csv
import datetime
import os
import openpyxl
from openpyxl.styles import Alignment, Font
from create_srg_export import COLUMN_MAPPINGS


SSG_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
RULES_JSON = os.path.join(SSG_ROOT, "build", "rule_dirs.json")
BUILD_CONFIG = os.path.join(SSG_ROOT, "build", "build_config.yml")
OUTPUT = os.path.join(SSG_ROOT, 'build',
                      f'{datetime.datetime.now().strftime("%s")}_stig_export.xlsx')

MICRO_COLUMN_SIZE = 8
SMALL_COLUMN_SIZE = 17
MEDIUM_COLUMN_SIZE = SMALL_COLUMN_SIZE*1.5
BIG_COLUMN_SIZE = SMALL_COLUMN_SIZE*3
HUGE_COLUMN_SIZE = SMALL_COLUMN_SIZE*4
COLUMN_SIZES = {
    'IA Control': SMALL_COLUMN_SIZE,
    'CCI': SMALL_COLUMN_SIZE,
    'SRGID': SMALL_COLUMN_SIZE,
    'STIGID': MICRO_COLUMN_SIZE,
    'SRG Requirement': HUGE_COLUMN_SIZE,
    'Requirement': HUGE_COLUMN_SIZE,
    'SRG VulDiscussion': HUGE_COLUMN_SIZE,
    'Vul Discussion': HUGE_COLUMN_SIZE,
    'Status': MEDIUM_COLUMN_SIZE,
    'SRG Check': HUGE_COLUMN_SIZE,
    'Check': HUGE_COLUMN_SIZE,
    'SRG Fix': MICRO_COLUMN_SIZE,
    'Fix': HUGE_COLUMN_SIZE,
    'Severity': MICRO_COLUMN_SIZE,
    'Mitigation': BIG_COLUMN_SIZE,
    'Artifact Description': BIG_COLUMN_SIZE,
    'Status Justification': BIG_COLUMN_SIZE
}

def setup_sheet(sheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
    for column, header in COLUMN_MAPPINGS.items():
        sheet.column_dimensions[f'{column}'].width = COLUMN_SIZES[header]


def setup_headers(sheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
    for column, header in COLUMN_MAPPINGS.items():
        sheet[f'{column}1'] = header
    for cell in list(sheet.iter_rows(max_row=1))[0]:
        cell.font = Font(bold=True, name='Calibri')


def format_cells(sheet: openpyxl.worksheet.worksheet.Worksheet):
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')


def setup_row(sheet: openpyxl.worksheet.worksheet.Worksheet, row: dict, row_num: int) -> None:
    for column, header in COLUMN_MAPPINGS.items():
        sheet[f'{column}{row_num}'] = row[header]
    sheet.row_dimensions[row_num].height = 130

    # freeze header row represented by A1
    # A2 is required because it freezes everything before it
    # in this case we only want A1 row to be frozen
    sheet.freeze_panes = "A2"


def handle_dict(data: list, output_path: str, sheet_name: str) -> None:
    """
    Given a dict with the fields for the srg export, create a formatted XLSX file
    """
    xlsx = openpyxl.Workbook()
    sheet = xlsx.active
    sheet.name = sheet_name
    setup_headers(sheet)
    setup_sheet(sheet)
    row_num = 2
    for row in data:
        setup_row(sheet, row, row_num)
        row_num += 1
    format_cells(sheet)
    xlsx.save(output_path)
