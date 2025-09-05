#!/usr/bin/python3
from __future__ import annotations

import argparse
import json
import os
import yaml

from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import load_workbook

from ssg.rule_yaml import find_section_lines, get_yaml_contents

SSG_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
RULES_JSON = os.path.join(SSG_ROOT, "build", "rule_dirs.json")
SEVERITY = {'CAT III': 'low', 'CAT II': 'medium', 'CAT I': 'high'}

# The start row is 2, to avoid importing the header
START_ROW = 2


class Row:
    row_id = 0
    IA_Control = ""
    CCI = ""
    SRGID = ""
    STIGID = ""
    SRG_Requirement = ""
    Requirement = ""
    SRG_VulDiscussion = ""
    Vul_Discussion = ""
    Status = ""
    SRG_Check = ""
    Check = ""
    SRG_Fix = ""
    Fix = ""
    Severity = ""
    Mitigation = ""
    Artifact_Description = ""
    Status_Justification = ""

    @staticmethod
    def from_row(sheet: Worksheet, row_number: int, full_name: str, changed_name: str) -> Row:
        self = Row()
        self.row_id = row_number
        self.IA_Control = sheet[f'A{row_number}'].value
        self.CCI = sheet[f'B{row_number}'].value
        self.SRGID = sheet[f'C{row_number}'].value
        self.STIGID = sheet[f'D{row_number}'].value
        self.SRG_Requirement = sheet[f'E{row_number}'].value
        self.Requirement = sheet[f'F{row_number}'].value.replace(full_name, changed_name)
        self.SRG_VulDiscussion = sheet[f'G{row_number}'].value
        self.Vul_Discussion = sheet[f'H{row_number}'].value
        self.Status = sheet[f'I{row_number}'].value
        self.SRG_Check = sheet[f'J{row_number}'].value
        self.Check = sheet[f'K{row_number}'].value
        self.SRG_Fix = sheet[f'L{row_number}'].value
        self.Fix = sheet[f'M{row_number}'].value
        self.Severity = sheet[f'N{row_number}'].value
        self.Mitigation = sheet[f'O{row_number}'].value
        self.Artifact_Description = sheet[f'P{row_number}'].value
        self.Artifact_Description = sheet[f'Q{row_number}'].value
        return self

    def __str__(self) -> str:
        return f'<Row {self.row_id}>'


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument('--changed', '-c', help="The file with changes, in xlsx", required=True)
    parser.add_argument('--current', '-b', help="The latest file from CaC, in xlsx", required=True)
    parser.add_argument('--changed-name', '-n', type=str, action="store",
                        help="The name that DISA uses for the producut. Defaults to RHEL 9",
                        default="RHEL 9")
    parser.add_argument("-r", "--root", type=str, action="store", default=SSG_ROOT,
                        help=f"Path to SSG root directory (defaults to {SSG_ROOT})")
    parser.add_argument("-j", "--json", type=str, action="store", default=RULES_JSON,
                        help=f"Path to the rules_dir.json (defaults to {RULES_JSON})")
    parser.add_argument("-p", "--product", type=str, action="store", required=True,
                        help="What product product id to use.")
    parser.add_argument("-e", '--end-row', type=int, action="store", default=600,
                        help="What row to end on, defaults to 600")
    return parser.parse_args()


def get_stigid_set(sheet: Worksheet, end_row: int) -> set[str]:
    result = set()
    start_row = 2
    for i in range(start_row, end_row):
        cci_raw = sheet[f'D{i}'].value

        if cci_raw is None or cci_raw == "":
            continue
        result.add(cci_raw)
    return result


def get_cce_dict_to_row_dict(sheet: Worksheet, full_name: str, changed_name: str,
                             end_row: int) -> dict:
    result = dict()
    for i in range(START_ROW, end_row):
        cci_raw = sheet[f'D{i}'].value
        if cci_raw is None or cci_raw == "":
            continue
        cci = cci_raw.strip()
        result[cci] = Row.from_row(sheet, i, full_name, changed_name)

    return result


def get_cce_dict(data: dict, product: str) -> dict:
    results = dict()
    for rule_id in data.keys():
        rule = data[rule_id]
        cce_key = f'cce@{product}' in rule['identifiers']
        if product in rule['products'] and cce_key:
            results[rule['identifiers'][f'cce@{product}']] = rule_id

    return results


def get_rule_dir_json(path: str) -> dict:
    with open(path, 'r') as f:
        return json.load(f)


def get_full_name(root_dir: str, product: str) -> str:
    product_yml_path = os.path.join(root_dir, 'products', product, 'product.yml')
    with open(product_yml_path, 'r') as f:
        data = yaml.load(f, Loader=yaml.SafeLoader)
        return data['full_name']


def get_cac_status(disa: str) -> str:
    return SEVERITY.get(disa, 'Unknown')


def replace_yaml_key(key: str, replacement: str, rule_dir: dict) -> None:
    path_dir = rule_dir['dir']
    path = os.path.join(path_dir, 'rule.yml')
    lines = get_yaml_contents(rule_dir)
    section_ranges = find_section_lines(lines.contents, key)
    replacement_line = f"{key}: {replacement}"
    if section_ranges:
        result = lines.contents[:section_ranges[0].start]
        result = (*result, replacement_line)
        end_line = section_ranges[0].end
        for line in lines.contents[end_line:]:
            result = (*result, line)
    else:
        result = lines.contents
        result = (*result, replacement_line)

    with open(path, 'w') as f:
        for line in result:
            f.write(line.rstrip())
            f.write('\n')


def replace_yaml_section(section: str, replacement: str, rule_dir: dict) -> None:
    path_dir = rule_dir['dir']
    path = os.path.join(path_dir, 'rule.yml')
    lines = get_yaml_contents(rule_dir)
    replacement = replacement.replace('RHEL 9', '{{{ full_name }}}')
    section_ranges = find_section_lines(lines.contents, section)
    if section_ranges:
        result = lines.contents[:section_ranges[0].start]
        result = (*result, f"{section}: |-")
        result = add_replacement_to_result(replacement, result)
        end_line = section_ranges[0].end
        for line in lines.contents[end_line:]:
            result = (*result, line)
    else:
        result = lines.contents
        result = (*result, f"\n{section}: |-")
        result = add_replacement_to_result(replacement, result)

    with open(path, 'w') as f:
        for line in result:
            f.write(line.rstrip())
            f.write('\n')


def add_replacement_to_result(replacement, result):
    for line in replacement.split("\n"):
        if line != "":
            result = (*result, f"    {line}",)
        else:
            result = (*result, "")
    return result


def fix_cac_cells(content: str, full_name: str, changed_name: str) -> str:
    if content:
        return content.replace(full_name, changed_name)
    return ""


def get_common_set(changed_sheet: Worksheet, current_sheet: Worksheet, end_row: int):
    changed_set = get_stigid_set(changed_sheet, end_row)
    current_set = get_stigid_set(current_sheet, end_row)
    common_set = current_set - (current_set - changed_set)
    return common_set


def update_severity(changed, current, rule_dir_json):
    if changed.Severity != current.Severity and changed.Severity is not None and \
            current.Severity is not None:
        cac_severity = get_cac_status(changed.Severity)
        replace_yaml_key('severity', cac_severity, rule_dir_json)


def update_row(changed: str, current: str, rule_dir_json: dict, section: str):
    if changed != current and changed:
        replace_yaml_section(section, changed, rule_dir_json)


def main() -> None:
    args = parse_args()
    full_name = get_full_name(args.root, args.product)
    changed_wb = load_workbook(args.changed)
    current_wb = load_workbook(args.current)
    current_sheet = current_wb['Sheet']
    changed_sheet = changed_wb['Sheet']
    common_set = get_common_set(changed_sheet, current_sheet, args.end_row)

    cac_cce_dict = get_cce_dict_to_row_dict(current_sheet, full_name, args.changed_name,
                                            args.end_row)
    disa_cce_dict = get_cce_dict_to_row_dict(changed_sheet, full_name, args.changed_name,
                                             args.end_row)

    rule_dir_json = get_rule_dir_json(args.json)
    cce_rule_id_dict = get_cce_dict(rule_dir_json, args.product)

    for cce in common_set:
        changed = disa_cce_dict[cce]
        current = cac_cce_dict[cce]
        rule_id = cce_rule_id_dict[cce]
        rule_obj = rule_dir_json[rule_id]

        update_row(changed.Requirement, current.Requirement, rule_obj, 'srg_requirement')

        update_row(changed.Vul_Discussion, current.Vul_Discussion, rule_obj, 'vuldiscussion')

        cleand_current_check = fix_cac_cells(current.Check, full_name, args.changed_name)
        update_row(changed.Check, cleand_current_check, rule_obj, 'checktext')

        cleand_current_fix = fix_cac_cells(current.Fix, full_name, args.changed_name)
        update_row(changed.Fix, cleand_current_fix, rule_obj, 'fixtext')

        update_severity(changed, current, rule_obj)


if __name__ == "__main__":
    main()
